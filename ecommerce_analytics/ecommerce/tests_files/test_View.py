from decimal import Decimal  
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase
from django.contrib.auth import get_user_model
from ..models import Customer, Order, OrderItem, Product, Inventory, Category  
from django.utils import timezone  
class SalesDataViewTest(APITestCase):
    def setUp(self):
        User = get_user_model()
         # Create a test user and log them in
        self.user = User.objects.create_user(username='testuser', password='testpass')
        self.client.login(username='testuser', password='testpass')
        # Create test data
        self.category = Category.objects.create(name="Electronics")
        self.product = Product.objects.create(
            name="Laptop",
            description="A powerful laptop.",
            price=Decimal('1500.00'),
            SKU="ABC123",
            category=self.category
        )
        self.inventory = Inventory.objects.create(product=self.product, quantity=10, last_restocked_date=timezone.now())
        self.customer = Customer.objects.create(name="John Doe", email="john@example.com", country="US")
        self.order = Order.objects.create(customer=self.customer, total_amount=Decimal('1500.00'))
        self.order_item = OrderItem.objects.create(order=self.order, product=self.product, quantity=1, price_at_time_of_order=Decimal('1500.00'))

    def test_get_sales_data_success(self):
        # Test successful retrieval of sales data
        url = reverse('sales-data')
        response = self.client.get(url, {'start_date': '2024-01-01', 'end_date': '2024-12-31'})
        self.assertEqual(response.status_code, status.HTTP_200_OK) # Check for 200 OK response
        revenue_data = response.data
        self.assertTrue(len(revenue_data) > 0)  # Ensure there is data returned
        for entry in revenue_data:
            self.assertIn('revenue', entry)  # Check if 'revenue' field is present
            self.assertIsInstance(entry['revenue'], Decimal)  # Check if revenue is of type Decimal

    def test_get_sales_data_missing_dates(self):
        # Test response when dates are missing
        self.client.login(username='testuser', password='testpass')
        url = reverse('sales-data')
        response = self.client.get(url) # No date parameters provided   
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST) # Expect a 400 Bad Request
        self.assertEqual(response.data, {"error": "Please provide a start and end date."}) # Check error message

from django.contrib.auth.models import User
from django.utils.timezone import now

class InventoryUpdateViewTest(APITestCase):
    def setUp(self):
        # Create test data for inventory update
        self.category = Category.objects.create(name="Electronics")
        self.product = Product.objects.create(name="Test Product", price=100.00, SKU="PROD123", category=self.category)
        self.inventory = Inventory.objects.create(product=self.product, quantity=10, last_restocked_date=now())
        self.url = reverse('inventory-update', kwargs={'pk': self.inventory.pk})

    def test_update_inventory_success(self):
       # Test successful inventory update
        data = {
            'quantity': 20,
            'last_restocked_date': now(),
            'product': self.product.id
        }
        response = self.client.post(self.url, data, format='json') # Send POST request to update inventory
        self.assertEqual(response.status_code, status.HTTP_200_OK)  # Check for successful response
        self.inventory.refresh_from_db() # Refresh from the database to get updated values
        self.assertEqual(self.inventory.quantity, 20)  # Check if quantity has been updated

    def test_update_inventory_not_found(self):
        # Test response for inventory not found 
        invalid_url = reverse('inventory-update', kwargs={'pk': 9999}) # Invalid primary key
        data = {
            'quantity': 20,
            'last_restocked_date': now(),
            'product': self.product.id
        }
        response = self.client.post(invalid_url, data, format='json')  # Attempt to update with invalid URL
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

# Test cases for Export Sales Report View

from openpyxl import load_workbook
from io import BytesIO
from django.test import TestCase

class ExportSalesReportViewTest(TestCase):

    def setUp(self):
        self.url = reverse('export-sales') # URL for exporting sales report

    def test_export_sales_report_success(self):
        # Test successful export of sales report
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK) # Check for successful response
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('attachment; filename="sales_report.xlsx"', response['Content-Disposition'])
        #Load the workbook and check its contents
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Monthly Sales Report")
        header = [cell.value for cell in ws[1]]
        self.assertEqual(header, ['Category', 'Revenue'])

# Test cases for Customer Info View
from ..serializers import CustomerSerializer
class CustomerInfoViewTest(APITestCase):
    def setUp(self):
        self.customer = Customer.objects.create(
            name='John Doe',
            email='johndoe@example.com',
            country='India',
            registration_date='2024-10-10T00:00:00Z' 
        )
        self.url = reverse('customer-info', kwargs={'pk': self.customer.pk})

    def test_get_customer_info_success(self):
        # Test successful retrieval of customer info
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        expected_data = CustomerSerializer(self.customer).data
        expected_data['registration_date'] = '2024-10-10T00:00:00Z'
        self.assertEqual(response.json(), expected_data)

    def test_get_customer_info_not_found(self):
        # Test response when customer is not found
        response = self.client.get(reverse('customer-info', kwargs={'pk': 999})) 
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)